from openpyxl import load_workbook

# Open the file with data on population
fin = open('statepopulation.txt', 'r', encoding='utf8')
statenames = dict()
for row in fin:
    row = row.split(",")
    statenames[row[0].upper()] = int(row[1].strip())
fin.close()

s = ["an", "ap", "ar", "as", "bh", "ch", "cg", "dd", "ddd", "dl", "gj", \
     "hr", "hp", "jk", "jh", "ka", "kl", "ld", "mp", "mh", "ml", "mz", \
     "mn", "od", "py", "pb", "rj", "sk", "tn", "tr", "up", "uk", "wb", \
     "na", "ga", "in"]
print(len(s))

for state in s:
    # Bilingual speakers
    wb = load_workbook('bi' + state + '.xlsx')
    s = wb['Sheet1'] # The data is in the first sheet.

    lang2 = dict() # Store data on second language speakers.
    lang3 = dict() # Store data on third language speakers.

    ROW_START = 7 # Data starts from this row.
    ROW_END = s.max_row # Data ends on this row.

    # Second language speakers 
    for i in range(ROW_START, ROW_END): # Data for the state is between rows 7 and 24767.
        # All languages, as opposed to dialects, codes end in three zeros. 
        if str(s['H'+str(i)].value)[-3:] == "000":  # On spotting a valid language code
            # Column I = Language name; Column J = Speakers 
            lang2[s['I'+str(i)].value] = lang2.get(s['I'+str(i)].value, 0) + int(s['J'+str(i)].value) 

    # Third language speakers 
    for i in range(ROW_START, ROW_END): # Data for the state is between rows 7 and 24767.
        # All languages, as opposed to dialects, codes end in three zeros. 
        if str(s['M'+str(i)].value)[-3:] == "000":
            # Column I = Language name; Column J = Speakers 
            lang3[s['N'+str(i)].value] = lang3.get(s['N'+str(i)].value, 0) + int(s['O'+str(i)].value)


    def plottable (lang_data, label):
        state = (s["B7"].value).strip()
        print(state)
        print(f'{"№":>3} {label:30} {"LEARNERS":>30} {"% OF POPULATION":>31}')
        count = 1 
        for k, v in sorted(lang_data.items(), key=lambda x:x[1], reverse=True):
            if v > 10000:
                k = k.strip()[:1] + k.strip()[1:].lower()
                print(f'{count:3} {k:30} {v:30} {(v/statenames[state]) * 100:30.2f}%')
                count += 1

    plottable(lang2, "SECOND LANGUAGE")
    plottable(lang3, "THIRD LANGUAGE")
