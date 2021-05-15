"""
Find the most popular mother tongues in each state or union territory.
"""

from openpyxl import load_workbook
import re

# Remove language codes
p = re.compile('[A-Z]+')

# Open the file with data on population
fin = open('statepopulation.txt', 'r', encoding='utf8')
statenames = dict()
for row in fin:
    row = row.split(",")
    statenames[row[0].upper()] = int(row[1].strip())
fin.close()

# Open the file with data on mother tongue
wb = load_workbook(filename='india.xlsx')
s = wb['Sheet1'] # The data is in the first sheet.
# Find out how many rows in each state, union territory, or the
# entire country occupies. 
states = dict()
for state in range(7, s.max_row):
    states[s['E'+str(state)].value] = states.get(s['E'+str(state)].value, 0) + 1

MIN_ROW = 7
MAX_ROW = MIN_ROW
for state in states.keys():
    MAX_ROW += states[state]
    languages = dict()
    for language in range(MIN_ROW, MAX_ROW): # Data for the state is between rows 7 and 307.
        if str(s['F'+str(language)].value)[-3:] == "000":  # On spotting a valid language code
            languages[str(s['G'+str(language)].value)[2:]] = s['H'+str(language)].value # Store the language name as key and speakers as value 

    print(f'{"№":>3} {"LANGUAGE":30} {"NATIVE SPEAKERS":>30} {"% OF POPULATION":>31}')
    count = 1
    for k, v in sorted(languages.items(), key=lambda x:x[1], reverse=True):
        if ((v/statenames[state]) * 100) > 1.0:
            k = p.search(k)
            k = k.group()
            k = k.strip()[:1] + k.strip()[1:].lower()
            print(f'{count:3} {k:30} {v:30} {(v/statenames[state]) * 100:30.2f}%')
            count += 1

    MIN_ROW = MAX_ROW

wb.close()
